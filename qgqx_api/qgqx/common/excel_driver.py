import openpyxl

from config import *


class ExcelDriver:

    # 读取excel的测试用例数据
    @staticmethod
    def read_excel(file_path=EXCEL_PATH, sheet_name=SHEET_NAME):
        # 打开现有的excel，如果没有新建一个
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # 选择指定的工作表/没有就创建
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)

        # 将excel数据存储为字典
        excel_data = []

        # 获取列名，以列表形式存储，获取第二行的数据
        headers = []
        for cell in worksheet[2]:
            headers.append(cell.value)

        # 获取每一列的数据
        for row in worksheet.iter_rows(min_row=3, values_only=True):  # 最小从第3行开始，获取值
            current_data = dict(zip(headers, row))
            # 判断是否需要执行
            if current_data["is_true"] == True:
                excel_data.append(current_data)
        # 关闭工作的excel
        workbook.close()

        return excel_data

    # 将执行结果写入excel
    @staticmethod
    def write_excel(file_path=EXCEL_PATH, sheet_name=SHEET_NAME, row=None, column=11, value=None):
        # 打开现有的excel，如果没有新建一个
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # 选择指定的工作表/没有就创建
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)

        # 写入指定的行和列
        worksheet.cell(row=row, column=column).value = value
        # 保存文件
        workbook.save(file_path)

# if __name__ == '__main__':
#     data = ExcelDriver.read_excel()
#     print(data)
#
#
# if __name__ == '__main__':
#     data = ExcelDriver.write_excel()
#     print(data)
